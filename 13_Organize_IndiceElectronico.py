import os
import pandas as pd
from openpyxl import load_workbook

def procesar_archivo_excel(archivo_excel_entrada, ruta_busqueda, prefijo):
    # Leer el archivo Excel de entrada sin encabezados
    df = pd.read_excel(archivo_excel_entrada, header=None)
    
    # Iterar sobre las filas desde la fila 2
    for index, row in df.iterrows():
        try:
            # Usar índices de columna en lugar de nombres
            valor_columna_c = str(row[1]).replace('-', '')  # Índice 1 para la columna B
            radicado = f"{prefijo}{valor_columna_c}00"
          
            demandante = str(row[4])  # Índice 4 para la columna E
            demandado = str(row[5])  # Índice 5 para la columna F
        except IndexError as e:
            print(f"Error: Índice no encontrado: {e}")
            continue
        print("=> Buscando radicado:" + radicado)
        # Buscar recursivamente en la ruta dada si hay una carpeta que comience con el radicado
        for root, dirs, files in os.walk(ruta_busqueda):
            for dir_name in dirs:
                if dir_name.startswith(radicado):
                    ruta_carpeta_encontrada = os.path.join(root, dir_name)
                    
                    # Buscar dentro de la carpeta el archivo de Excel donde se va a modificar la info
                    for subdir_root, subdirs, subfiles in os.walk(ruta_carpeta_encontrada):
                        for file in subfiles:
                            if file.startswith('00IndiceElectronicoC0'):
                                archivo_excel_salida = os.path.join(subdir_root, file)
                                
                                # Cargar el archivo Excel y modificar las celdas B5, B6, B7
                                modificar_archivo_excel(archivo_excel_salida, radicado, demandante, demandado)
                                print(f"Modificado archivo: {archivo_excel_salida}")

def modificar_archivo_excel(archivo, radicado, demandante, demandado):
    # Cargar el archivo de Excel con openpyxl
    wb = load_workbook(archivo, keep_vba=True)
    ws = wb.active
    
    # Modificar las celdas B5, B6 y B7
    ws['B5'] = radicado
    ws['B6'] = demandado
    ws['B7'] = demandante
    
    print("Indice Organizado")
    # Guardar el archivo
    wb.save(archivo)


# Ejemplo de uso
archivo_excel_entrada = './BaseDatos/Archivo.xlsx'  # Reemplaza con el nombre del archivo de entrada
ruta_busqueda = r'D:\OneDrive - Consejo Superior de la Judicatura\01. EXPEDIENTES DE PROCESOS JUDICIALES'  # Reemplaza con la ruta de búsqueda
prefijo = '056314089001'  # Reemplaza con el prefijo que desees

procesar_archivo_excel(archivo_excel_entrada, ruta_busqueda, prefijo)




