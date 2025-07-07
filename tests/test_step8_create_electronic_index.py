import os
import tempfile
import unittest
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from organizer import step8_create_electronic_index as sei


class TestElectronicIndex(unittest.TestCase):
    @patch(
        "organizer.step8_create_electronic_index.os.listdir", return_value=[]
    )
    def test_get_empty_folders(self, _):
        res = sei.get_empty_folders("path")
        self.assertEqual(len(res), 1)
        self.assertEqual(res[0]["Causa del problema"], "Carpeta vacía")

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["a.txt"],
    )
    @patch(
        "organizer.step8_create_electronic_index.os.path.getsize",
        return_value=0,
    )
    @patch(
        "organizer.step8_create_electronic_index.os.path.isfile",
        return_value=True,
    )
    def test_get_empty_files(self, _, mock_size, mock_list):
        res = sei.get_empty_files("path")
        self.assertTrue(res)
        self.assertEqual(res[0]["Archivo Inválido"], "a.txt")

    @patch(
        "organizer.step8_create_electronic_index.os.path.getsize",
        return_value=512,
    )
    def test_format_file_size_bytes(self, mock_size):
        self.assertEqual(sei.format_file_size("x"), "512 B")

    @patch(
        "organizer.step8_create_electronic_index.os.path.getsize",
        return_value=2048,
    )
    def test_format_file_size_kb(self, _):
        self.assertIn("KB", sei.format_file_size("x"))

    @patch(
        "organizer.step8_create_electronic_index.os.path.getsize",
        return_value=1048576,
    )
    def test_format_file_size_mb(self, _):
        self.assertIn("MB", sei.format_file_size("x"))

    @patch(
        "organizer.step8_create_electronic_index.os.path.getsize",
        return_value=5 * 1024**3,
    )
    def test_format_file_size_gb(self, _):
        self.assertIn("GB", sei.format_file_size("x"))

    def test_valid_document(self):
        self.assertTrue(sei.valid_document("01doc.pdf"))
        self.assertFalse(sei.valid_document(".hidden"))
        self.assertFalse(sei.valid_document("desktop.ini"))
        self.assertFalse(sei.valid_document("00IndiceElectronicoC01.xlsm"))

    def test_filter_target_folders(self):
        dirs = ["01PrimeraInstancia", "x", "02SecInst"]
        out = sei.filter_target_folders(dirs)
        self.assertEqual(out, ["01PrimeraInstancia", "02SecInst"])

    @patch(
        "organizer.step8_create_electronic_index.fitz.open",
        return_value=MagicMock(page_count=5),
    )
    @patch(
        "organizer.step8_create_electronic_index.os.path.getsize",
        return_value=2048,
    )
    @patch(
        "organizer.step8_create_electronic_index.os.path.getctime",
        return_value=0,
    )
    def test_get_file_info_pdf(self, mock_ctime, mock_size, mock_fitz):
        f = "01file.PDF"
        info = sei.get_file_info(f)
        self.assertEqual(info["file_extension"], "PDF")
        self.assertEqual(info["page_count"], 5)
        self.assertEqual(info["file_number"], 1)
        self.assertIn("KB", info["file_size"])
        self.assertEqual(info["name"], "01file.PDF")

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["01a.xlsx"],
    )
    @patch(
        "organizer.step8_create_electronic_index.load_workbook",
        return_value=None,
    )
    def test_validate_excels_valid(self, mock_wb, mock_list):
        self.assertIsNone(sei.validate_excels_in_folder("p"))

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["01a.xlsm"],
    )
    @patch(
        "organizer.step8_create_electronic_index.load_workbook",
        side_effect=Exception("err"),
    )
    def test_validate_excels_invalid(self, mock_wb, _):
        res = sei.validate_excels_in_folder("p")
        self.assertEqual(res[0]["Causa del problema"], "err")

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["01a.docx"],
    )
    @patch(
        "organizer.step8_create_electronic_index.Document", return_value=None
    )
    def test_validate_word_valid(self, mock_doc, _):
        self.assertIsNone(sei.validate_word_docs_in_folder("p"))

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["01a.docx"],
    )
    @patch(
        "organizer.step8_create_electronic_index.Document",
        side_effect=Exception("bad"),
    )
    def test_validate_word_invalid(self, mock_doc, _):
        res = sei.validate_word_docs_in_folder("p")
        self.assertEqual(res[0]["Causa del problema"], "bad")

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["01a.pdf"],
    )
    @patch(
        "organizer.step8_create_electronic_index.fitz.open", return_value=None
    )
    def test_validate_pdf_valid(self, mock_fitz, _):
        self.assertIsNone(sei.validate_pdfs_in_folder("p"))

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["01a.pdf"],
    )
    @patch(
        "organizer.step8_create_electronic_index.fitz.open",
        side_effect=Exception("errpdf"),
    )
    def test_validate_pdf_invalid(self, mock_fitz, _):
        res = sei.validate_pdfs_in_folder("p")
        self.assertEqual(res[0]["Causa del problema"], "errpdf")

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["01a.pdf", "Zcontrol.xlsx"],
    )
    def test_validate_prefix_ok(self, mock_list):
        self.assertIsNone(sei.validate_files_with_numeric_prefix("p"))

    @patch(
        "organizer.step8_create_electronic_index.os.listdir",
        return_value=["a.pdf"],
    )
    def test_validate_prefix_fail(self, mock_list):
        res = sei.validate_files_with_numeric_prefix("p")
        self.assertEqual(
            res[0]["Causa del problema"], "Sin prefijo numérico de 2 dígitos"
        )

    @patch(
        "organizer.step8_create_electronic_index.os.path.exists",
        return_value=True,
    )
    def test_validate_if_file_exists_true(self, mock_exists):
        self.assertTrue(sei.validate_if_file_exists("p", "1"))

    @patch(
        "organizer.step8_create_electronic_index.os.path.exists",
        return_value=False,
    )
    def test_validate_if_file_exists_false(self, mock_exists):
        self.assertFalse(sei.validate_if_file_exists("p", "2"))

    @patch(
        "organizer.step8_create_electronic_index.os.path.isfile",
        return_value=False,
    )
    def test_buscar_no_file(self, mock_isfile):
        res = sei.buscar_radicado_en_base_de_datos("x")
        self.assertEqual(res, [])

    @patch("organizer.step8_create_electronic_index.pd.read_excel")
    @patch(
        "organizer.step8_create_electronic_index.os.path.isfile",
        return_value=True,
    )
    def test_buscar_found(self, mock_isfile, mock_read):
        # create df with >=6 cols
        data = {i: ["", ""] for i in range(6)}
        data[1] = ["xyz05380123456789012345678", "other"]
        data[4] = ["P", "?"]
        data[5] = ["Q", "?"]
        df = pd.DataFrame(data)
        mock_read.return_value = df
        res = sei.buscar_radicado_en_base_de_datos("05380123456789012345678")
        self.assertEqual(res[0][4], "P")
        self.assertEqual(res[0][5], "Q")

    def test_get_radicado_number(self):
        path = "/x/05380123456789012345678/test"
        self.assertEqual(
            sei.get_radicado_number(path), "05380123456789012345678"
        )
        self.assertEqual(sei.get_radicado_number("no_match"), "")

    def test_delete_rows_from_excel(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb = Workbook()
        ws = wb.active
        for i in range(1, 21):
            ws.cell(row=i, column=1, value=i)
        wb.save(tmp.name)
        tmp.close()
        sei.delete_rows_from_excel(tmp.name, 5, 10)
        wb2 = load_workbook(tmp.name)
        # rows deleted: 6 rows => 20-6 = 14
        self.assertEqual(wb2.active.max_row, 14)
        wb2.close()
        os.unlink(tmp.name)

    def test_apply_border_to_row(self):
        wb = Workbook()
        ws = wb.active
        sei.apply_border_to_row(ws, 3)
        cell = ws.cell(row=3, column=1)
        self.assertTrue(cell.border.top.style)
        self.assertEqual(cell.alignment.horizontal, "center")

    @patch("organizer.step8_create_electronic_index.shutil.copy")
    @patch("organizer.step8_create_electronic_index.delete_rows_from_excel")
    @patch("organizer.step8_create_electronic_index.load_workbook")
    @patch("organizer.step8_create_electronic_index.insert_rows")
    @patch(
        "organizer.step8_create_electronic_index.buscar_radicado_en_base_de_datos"
    )
    def test_generate_index_file(
        self, mock_bs, mock_ins, mock_load, mock_del, mock_copy
    ):
        tmp = tempfile.NamedTemporaryFile(delete=False)
        tmp.close()
        sei.config.TEMPLATE_FILE = tmp.name
        ws = MagicMock()
        wb = MagicMock(active=ws)
        mock_load.return_value = wb
        mock_bs.return_value = [{4: "X", 5: "Y"}]
        with tempfile.TemporaryDirectory() as td:
            result = sei.generate_index_file(td, "C01Test", "1", "rad")
            self.assertEqual(
                result["Causa del problema"], "Generado correctamente"
            )
            mock_copy.assert_called_once()
            mock_del.assert_called_once()
            mock_load.assert_called_once()
            mock_bs.assert_called_with("rad")
            mock_ins.assert_called_with(ws, [{4: "X", 5: "Y"}])
        os.unlink(tmp.name)


if __name__ == "__main__":
    unittest.main()
